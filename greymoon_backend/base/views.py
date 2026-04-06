import re
import io
import uuid
import requests
from datetime import datetime, timezone, timedelta
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import ServiceLead, ScrapeRun, ScrapedFbGroup
from .serializers import ServiceLeadSerializer, ScrapeRunSerializer
from .services.tasks import start_pipeline_thread
from .services.category_map import ALL_CATEGORIES, SERVICE_CATEGORY_MAP
from .services.location_resolver import resolve_location, LocationResolutionError
from .services.city_structure import US_CITY_STRUCTURE

PAGE_SIZE = 50

_STATE_NAME_TO_ABBREV = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
}


def _apply_lead_filters(leads, params):
    source = params.get("source")
    if source:
        leads = leads.filter(source=source.upper())

    service_category = params.get("service_category")
    if service_category:
        leads = leads.filter(service_category=service_category.lower())

    status = params.get("status")
    if status:
        leads = leads.filter(status=status.upper())

    min_score = params.get("min_score")
    if min_score:
        try:
            leads = leads.filter(score__gte=int(min_score))
        except ValueError:
            pass

    search = params.get("search")
    if search:
        from django.db.models import Q
        leads = leads.filter(
            Q(location__icontains=search) | Q(title__icontains=search)
        )

    if params.get("has_phone") in ("true", "1"):
        leads = leads.exclude(phone__isnull=True).exclude(phone="")
    if params.get("has_email") in ("true", "1"):
        leads = leads.exclude(email__isnull=True).exclude(email="")

    fb_group = params.get("fb_group")
    if fb_group:
        leads = leads.filter(fb_group_name__icontains=fb_group)

    date_from = params.get("date_from")
    if date_from:
        try:
            from django.utils.dateparse import parse_date
            d = parse_date(date_from)
            if d:
                leads = leads.filter(created_at__date__gte=d)
        except Exception:
            pass

    date_to = params.get("date_to")
    if date_to:
        try:
            from django.utils.dateparse import parse_date
            d = parse_date(date_to)
            if d:
                leads = leads.filter(created_at__date__lte=d)
        except Exception:
            pass

    return leads

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def manual_scrape(request):
    data = request.data

    sources = data.get("sources", ["craigslist"])
    valid_sources   = {"craigslist", "facebook", "indeed", "google"}
    invalid_sources = [s for s in sources if s not in valid_sources]
    if invalid_sources:
        return Response(
            {"error": f"Invalid sources: {invalid_sources}. Valid: {list(valid_sources)}"},
            status=400,
        )

    location = data.get("location") or {}
    location_type  = location.get("type", "").lower() if location else ""
    location_value = (location.get("value") or "").strip() if location else ""

    if location_type and location_type not in ("state", "city", "zip"):
        return Response({"error": "location.type must be 'state', 'city', or 'zip'."}, status=400)

    fb_group_urls_raw = data.get("fb_group_urls") or []
    if isinstance(fb_group_urls_raw, str):
        fb_group_urls_raw = [u.strip() for u in fb_group_urls_raw.split(",") if u.strip()]
    fb_group_urls = [u for u in fb_group_urls_raw if u.startswith("http")]

    fb_only = set(sources) == {"facebook"}

    if not location_value and not fb_group_urls:
        return Response({"error": "Provide a location or at least one Facebook group URL."}, status=400)
    if not location_value and "craigslist" in sources:
        return Response({"error": "location.value is required when scraping Craigslist."}, status=400)

    categories = data.get("categories", [])
    non_fb_sources = [s for s in sources if s != "facebook"]
    if non_fb_sources and not categories:
        return Response(
            {"error": f"At least one category required for {non_fb_sources}. Options: {ALL_CATEGORIES}"},
            status=400,
        )

    if categories:
        invalid_cats = [c for c in categories if c not in ALL_CATEGORIES]
        if invalid_cats:
            return Response(
                {"error": f"Invalid categories: {invalid_cats}. Valid: {ALL_CATEGORIES}"},
                status=400,
            )

    max_posts_per_group = int(data.get("max_posts_per_group", 50))
    max_posts_per_group = max(5, min(max_posts_per_group, 500))

    google_max_pages   = int(data.get("google_max_pages", 3))
    google_max_pages   = max(1, min(google_max_pages, 10))
    google_deep_scrape = bool(data.get("google_deep_scrape", True))

    # ── max_leads (0 = unlimited) ──────────────────────────────
    try:
        max_leads = int(data.get("max_leads", 0))
    except (TypeError, ValueError):
        max_leads = 0
    max_leads = max(0, max_leads)

    if location_value:
        try:
            location_data = resolve_location(location_type, location_value)
        except LocationResolutionError as e:
            return Response({"error": str(e)}, status=400)
        location_display = location_data["display"]
    else:
        location_data    = None
        location_display = "Facebook Groups"

    run = ScrapeRun.objects.create(
        run_id=str(uuid.uuid4()),
        status="RUNNING",
        location_type=location_type or "custom",
        location_value=location_value or "",
        location_display=location_display,
        max_posts_per_group=max_posts_per_group,
        categories=categories,
        sources=sources,
        current_stage="Starting",
        stage_detail="Pipeline initialising…",
        activity_log=[],
        google_max_pages=google_max_pages,
        google_deep_scrape=google_deep_scrape,
        max_leads=max_leads,
    )

    start_pipeline_thread(
        location_type=location_type or "custom",
        location_value=location_value or "",
        categories=categories,
        sources=sources,
        scrape_run_id=run.pk,
        max_posts_per_group=max_posts_per_group,
        fb_group_urls=fb_group_urls,
        google_max_pages=google_max_pages,
        google_deep_scrape=google_deep_scrape,
        max_leads=max_leads,
    )

    return Response({
        "message": "Scrape started successfully.",
        "run_id":  run.run_id,
        "location": location_display,
        "categories": categories,
        "sources": sources,
        "fb_groups": len(fb_group_urls),
        "max_leads": max_leads or "unlimited",
        "estimated_time": "2–10 minutes depending on sources and group size",
    }, status=202)


def _sweep_and_abort(apify_headers: dict, already_aborted: set) -> list[str]:
    newly = []
    for status_filter in ("RUNNING", "READY"):
        try:
            resp = requests.get(
                "https://api.apify.com/v2/actor-runs",
                headers=apify_headers,
                params={"status": status_filter, "limit": 100},
                timeout=15,
            )
            if resp.status_code != 200:
                continue
            for actor_run in resp.json().get("data", {}).get("items", []):
                aid = actor_run.get("id")
                if aid and aid not in already_aborted:
                    try:
                        ar = requests.post(
                            f"https://api.apify.com/v2/actor-runs/{aid}/abort",
                            headers=apify_headers,
                            timeout=10,
                        )
                        if ar.status_code in (200, 204):
                            newly.append(aid)
                            already_aborted.add(aid)
                    except Exception as e:
                        print(f"[Cancel] Could not abort actor {aid}: {e}")
        except Exception as e:
            print(f"[Cancel] Sweep ({status_filter}) error: {e}")
    return newly


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def cancel_scrape(request):
    run_id = request.data.get("run_id")
    if not run_id:
        return Response({"error": "run_id required"}, status=400)

    run = ScrapeRun.objects.filter(run_id=run_id).first()
    if not run:
        return Response({"error": "Run not found"}, status=404)

    ScrapeRun.objects.filter(run_id=run_id).update(cancel_requested=True)

    apify_headers  = {"Authorization": f"Bearer {settings.APIFY_TOKEN}"}
    aborted_set: set = set()

    for apify_id in (run.apify_run_ids or []):
        try:
            resp = requests.post(
                f"https://api.apify.com/v2/actor-runs/{apify_id}/abort",
                headers=apify_headers,
                timeout=10,
            )
            if resp.status_code in (200, 204):
                aborted_set.add(apify_id)
        except Exception as e:
            print(f"[Cancel] Abort failed for registered actor {apify_id}: {e}")

    _sweep_and_abort(apify_headers, aborted_set)

    import time as _time
    _time.sleep(3)
    _sweep_and_abort(apify_headers, aborted_set)

    run.refresh_from_db()
    detail = (
        f"Run cancelled — {run.leads_collected} lead(s) already saved to database."
        if run.leads_collected > 0
        else "Run cancelled before any leads were collected."
    )
    activity_log = run.activity_log or []
    activity_log.append({
        "ts":     datetime.now(timezone.utc).isoformat(),
        "stage":  "Stopped by user",
        "detail": detail,
        "level":  "warning",
    })
    run.status        = "PARTIAL" if run.leads_collected > 0 else "ABORTED"
    run.finished_at   = datetime.now(timezone.utc)
    run.current_stage = "Stopped by user"
    run.stage_detail  = detail
    run.activity_log  = activity_log
    run.save()

    return Response({
        "message":            f"Scrape {run_id} cancelled.",
        "apify_runs_aborted": len(aborted_set),
        "leads_saved_so_far": run.leads_collected,
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def scrape_status(request):
    run = ScrapeRun.objects.order_by("-created_at").first()
    if not run:
        return Response({"status": "IDLE"})

    source_stats = {}
    try:
        source_stats = run.source_stats or {}
    except AttributeError:
        pass

    if not source_stats:
        for entry in (run.activity_log or []):
            if entry.get("type") == "source_stats":
                src = entry.get("source")
                if src:
                    source_stats[src] = {
                        "saved":   entry.get("saved", 0),
                        "skipped": entry.get("skipped", 0),
                    }

    return Response({
        "run_id":          run.run_id,
        "status":          run.status,
        "location":        run.location_display,
        "categories":      run.categories,
        "sources":         run.sources,
        "leads_collected": run.leads_collected,
        "leads_skipped":   run.leads_skipped,
        "source_stats":    source_stats,
        "current_stage":   run.current_stage or "",
        "stage_detail":    run.stage_detail or "",
        "activity_log":    run.activity_log or [],
        "started_at":      run.created_at,
        "finished_at":     run.finished_at,
        "max_leads":       run.max_leads,
        "limit_stop":      run.limit_stop,
        "cancel_requested": run.cancel_requested,
        "is_stopping": (
                        run.status == "RUNNING" and (
                        run.cancel_requested or
                        (run.current_stage or "").lower().startswith("stopping")
    )
),

    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def scrape_history(request):
    runs = ScrapeRun.objects.order_by("-created_at")[:50]
    return Response([
        {
            "run_id":          r.run_id,
            "status":          r.status,
            "location":        r.location_display,
            "categories":      r.categories,
            "sources":         r.sources,
            "leads_collected": r.leads_collected,
            "leads_skipped":   r.leads_skipped,
            "started_at":      r.created_at,
            "finished_at":     r.finished_at,
            "max_leads":       r.max_leads,
            "limit_stop":      r.limit_stop,
        }
        for r in runs
    ])


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_fb_groups(request):
    urls_raw = request.data.get("group_urls") or []
    if isinstance(urls_raw, str):
        urls_raw = [u.strip() for u in re.split(r"[\n,]+", urls_raw) if u.strip()]

    valid_urls = [u for u in urls_raw if u.startswith("http")]
    if not valid_urls:
        return Response({"error": "No valid group URLs provided."}, status=400)

    added = []
    already_exists = []
    for url in valid_urls:
        url = url.rstrip("/") + "/"
        obj, created = ScrapedFbGroup.objects.get_or_create(
            group_url=url,
            defaults={"group_name": "", "post_count": 0},
        )
        if created:
            added.append(url)
        else:
            already_exists.append(url)

    return Response({
        "added":          len(added),
        "already_exists": len(already_exists),
        "total":          ScrapedFbGroup.objects.count(),
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_scraped_groups(request):
    def _display_name(group_name: str, group_url: str) -> str:
        if group_name and group_name.strip():
            return group_name.strip()
        m = re.search(r"/groups/([^/?#]+)", group_url or "")
        if m:
            slug = m.group(1)
            slug = re.sub(r"([a-z])([A-Z])", r"\1 \2", slug)
            slug = re.sub(r"[-_]", " ", slug)
            slug = re.sub(r"\s+", " ", slug).strip()
            return slug.title() if slug else group_url
        return group_url or "Unknown group"

    groups = list(ScrapedFbGroup.objects.all().values(
        "id", "group_url", "group_name", "post_count", "last_scraped", "first_scraped"
    ))
    for g in groups:
        g["group_name"] = _display_name(g["group_name"], g["group_url"])

    return Response({"groups": groups, "total": len(groups)})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def scrape_selected_groups(request):
    group_urls = request.data.get("group_urls") or []
    if isinstance(group_urls, str):
        group_urls = [u.strip() for u in re.split(r"[\n,]+", group_urls) if u.strip()]
    group_urls = [u for u in group_urls if u.startswith("http")]

    if not group_urls:
        return Response({"error": "No valid group URLs provided."}, status=400)

    max_posts_per_group = int(request.data.get("max_posts_per_group", 50))
    max_posts_per_group = max(5, min(max_posts_per_group, 500))

    try:
        max_leads = int(request.data.get("max_leads", 0))
    except (TypeError, ValueError):
        max_leads = 0
    max_leads = max(0, max_leads)

    run = ScrapeRun.objects.create(
        run_id=str(uuid.uuid4()),
        status="RUNNING",
        location_type="custom",
        location_value="",
        location_display=f"{len(group_urls)} Facebook group(s)",
        max_posts_per_group=max_posts_per_group,
        categories=[],
        sources=["facebook"],
        current_stage="Starting",
        stage_detail="Pipeline initialising…",
        activity_log=[],
        google_max_pages=3,
        google_deep_scrape=False,
        max_leads=max_leads,
    )

    start_pipeline_thread(
        location_type="custom",
        location_value="",
        categories=[],
        sources=["facebook"],
        scrape_run_id=run.pk,
        max_posts_per_group=max_posts_per_group,
        fb_group_urls=group_urls,
        max_leads=max_leads,
    )

    return Response({
        "message":   "Facebook scrape started.",
        "run_id":    run.run_id,
        "groups":    len(group_urls),
        "max_posts": max_posts_per_group,
        "max_leads": max_leads or "unlimited",
    }, status=202)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_group_leads(request):
    group_url = request.query_params.get("group_url", "")
    if not group_url:
        return Response({"error": "group_url required"}, status=400)
    leads       = ServiceLead.objects.filter(fb_group_url=group_url).order_by("-created_at")
    page_size   = min(200, int(request.query_params.get("page_size", 50)))
    page        = max(1, int(request.query_params.get("page", 1)))
    total       = leads.count()
    total_pages = max(1, (total + page_size - 1) // page_size)
    page        = min(page, total_pages)
    offset      = (page - 1) * page_size
    serializer  = ServiceLeadSerializer(leads[offset:offset + page_size], many=True)
    return Response({
        "results":     serializer.data,
        "total":       total,
        "page":        page,
        "total_pages": total_pages,
        "has_next":    page < total_pages,
    })


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_scraped_group(request):
    group_url = request.data.get("group_url", "")
    if not group_url:
        return Response({"error": "group_url required"}, status=400)
    deleted, _ = ScrapedFbGroup.objects.filter(group_url=group_url).delete()
    return Response({"deleted": deleted})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_services(request):
    leads = ServiceLead.objects.all()

    leads = _apply_lead_filters(leads, request.query_params)

    ordering_param = request.query_params.get("ordering", "-created_at")
    ALLOWED_ORDERING_FIELDS = {
        "datetime", "-datetime",
        "score", "-score",
        "created_at", "-created_at",
    }
    if ordering_param not in ALLOWED_ORDERING_FIELDS:
        ordering_param = "-created_at"

    if ordering_param == "-datetime":
        leads = leads.order_by("-created_at")
    elif ordering_param == "datetime":
        leads = leads.order_by("created_at")
    else:
        leads = leads.order_by(ordering_param, "-created_at")

    total = leads.count()

    try:
        page = max(1, int(request.query_params.get("page", 1)))
    except ValueError:
        page = 1

    try:
        page_size = min(200, max(1, int(request.query_params.get("page_size", PAGE_SIZE))))
    except ValueError:
        page_size = PAGE_SIZE

    total_pages = max(1, (total + page_size - 1) // page_size)
    page        = min(page, total_pages)
    offset      = (page - 1) * page_size
    page_leads  = leads[offset:offset + page_size]

    serializer = ServiceLeadSerializer(page_leads, many=True)
    return Response({
        "results":     serializer.data,
        "total":       total,
        "page":        page,
        "page_size":   page_size,
        "total_pages": total_pages,
        "has_next":    page < total_pages,
        "has_prev":    page > 1,
    })


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def update_lead_status(request, post_id):
    try:
        lead = ServiceLead.objects.get(post_id=post_id)
    except ServiceLead.DoesNotExist:
        return Response({"error": "Lead not found"}, status=404)

    new_status = request.data.get("status")
    if new_status not in dict(ServiceLead.STATUS_CHOICES):
        return Response({"error": "Invalid status"}, status=400)

    lead.status = new_status
    lead.save()
    return Response({"message": "Status updated", "status": new_status})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_leads(request):
    leads = ServiceLead.objects.all()

    leads = _apply_lead_filters(leads, request.query_params)
    leads = leads.order_by("-score", "-created_at")

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    HEADER_BG  = "1E293B"
    HEADER_FG  = "F8FAFC"
    CL_BG      = "FFF7ED"
    FB_BG      = "EEF2FF"
    GG_BG      = "F0F9FF"
    ALT_BG     = "F8FAFC"
    HIGH_SCORE = "D1FAE5"
    MED_SCORE  = "FEF9C3"
    BORDER_CLR = "E2E8F0"

    SOURCE_BG = {
        "CRAIGSLIST": CL_BG,
        "FACEBOOK":   FB_BG,
        "GOOGLE":     GG_BG,
    }
    SOURCE_LABEL = {
        "CRAIGSLIST": "CL",
        "FACEBOOK":   "FB",
        "GOOGLE":     "GG",
    }

    def thin_border():
        s = Side(style="thin", color=BORDER_CLR)
        return Border(left=s, right=s, top=s, bottom=s)

    def header_font():
        return Font(name="Arial", bold=True, color=HEADER_FG, size=9)

    def cell_font(bold=False, color="1E293B", size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=False)

    COLS = [
        ("Source",       "source",           8,   "center"),
        ("Score",        "score",            7,   "center"),
        ("Status",       "status",           11,  "center"),
        ("Title",        "title",            42,  "left"),
        ("Phone",        "phone",            16,  "left"),
        ("Email",        "email",            28,  "left"),
        ("Location",     "location",         20,  "left"),
        ("State",        "state",            8,   "center"),
        ("Category",     "service_category", 18,  "left"),
        ("FB Group",     "fb_group_name",    24,  "left"),
        ("URL",          "url",              40,  "left"),
        ("Posted",       "datetime",         14,  "center"),
        ("Scraped",      "created_at",       14,  "center"),
    ]

    ws.row_dimensions[1].height = 20
    for col_idx, (header, _, width, align) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font()
        cell.fill      = fill(HEADER_BG)
        cell.alignment = center() if align == "center" else left()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    for row_idx, lead in enumerate(leads, start=2):
        ws.row_dimensions[row_idx].height = 16
        src    = (lead.source or "").upper()
        row_bg = SOURCE_BG.get(src, ALT_BG if row_idx % 2 == 0 else "FFFFFF")

        for col_idx, (_, field, _, align) in enumerate(COLS, start=1):
            if field == "source":
                value = SOURCE_LABEL.get(src, src)
            elif field == "score":
                value = lead.score
            elif field == "status":
                value = lead.status
            elif field == "title":
                value = (lead.title or "")[:200]
            elif field == "phone":
                value = lead.phone or ""
            elif field == "email":
                value = lead.email or ""
            elif field == "location":
                value = lead.location or ""
            elif field == "state":
                value = lead.state or ""
            elif field == "service_category":
                value = lead.service_category or lead.category or ""
            elif field == "fb_group_name":
                value = lead.fb_group_name or ""
            elif field == "url":
                value = lead.url or ""
            elif field == "datetime":
                value = lead.datetime.strftime("%Y-%m-%d") if lead.datetime else ""
            elif field == "created_at":
                value = lead.created_at.strftime("%Y-%m-%d") if lead.created_at else ""
            else:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border()
            cell.alignment = center() if align == "center" else left()

            if field == "score":
                score_val = lead.score or 0
                if score_val >= 70:
                    cell.fill = fill(HIGH_SCORE)
                    cell.font = cell_font(bold=True, color="065F46")
                elif score_val >= 40:
                    cell.fill = fill(MED_SCORE)
                    cell.font = cell_font(bold=True, color="92400E")
                else:
                    cell.fill = fill(row_bg)
                    cell.font = cell_font(color="64748B")
            elif field == "phone" and lead.phone:
                cell.fill = fill(row_bg)
                cell.font = cell_font(bold=True, color="065F46")
            elif field == "email" and lead.email:
                cell.fill = fill(row_bg)
                cell.font = cell_font(bold=True, color="4338CA")
            elif field == "url" and lead.url:
                cell.fill      = fill(row_bg)
                cell.font      = Font(name="Arial", size=9, color="2563EB", underline="single")
                cell.hyperlink = lead.url
            else:
                cell.fill = fill(row_bg)
                cell.font = cell_font(
                    bold=(field == "title"),
                    color="1E293B" if field == "title" else "475569",
                )

    total_rows  = leads.count()
    summary_row = total_rows + 3
    ws.cell(row=summary_row,     column=1, value="Total leads exported:").font = cell_font(bold=True)
    ws.cell(row=summary_row,     column=2, value=total_rows).font              = cell_font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="Exported at:").font          = cell_font(bold=True)
    from django.utils import timezone as tz
    ws.cell(row=summary_row + 1, column=2,
            value=tz.now().strftime("%Y-%m-%d %H:%M UTC")).font = cell_font()

    filters_applied = []
    if request.query_params.get("source"):
        filters_applied.append(f"Source: {request.query_params['source']}")
    if request.query_params.get("date_from"):
        filters_applied.append(f"From: {request.query_params['date_from']}")
    if request.query_params.get("date_to"):
        filters_applied.append(f"To: {request.query_params['date_to']}")
    if request.query_params.get("status"):
        filters_applied.append(f"Status: {request.query_params['status']}")
    if request.query_params.get("min_score"):
        filters_applied.append(f"Min score: {request.query_params['min_score']}")
    if request.query_params.get("has_phone") in ("true", "1"):
        filters_applied.append("Has phone")
    if request.query_params.get("has_email") in ("true", "1"):
        filters_applied.append("Has email")
    if filters_applied:
        ws.cell(row=summary_row + 2, column=1, value="Filters:").font = cell_font(bold=True)
        ws.cell(row=summary_row + 2, column=2,
                value=" | ".join(filters_applied)).font = cell_font(color="475569")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from django.utils import timezone as tz2
    filename = f"leads_export_{tz2.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_cities(request):
    state_filter = request.query_params.get("state", "").title()
    cities_data  = []
    for state_name, state_info in US_CITY_STRUCTURE.items():
        if state_filter and state_name != state_filter:
            continue
        abbrev = _STATE_NAME_TO_ABBREV.get(state_name, state_name)
        for region in state_info["regions"]:
            cities_data.append({
                "code":    region["code"],
                "name":    region["name"],
                "state":   abbrev,
                "display": f"{region['name']}, {abbrev}",
            })
    return Response({"cities": cities_data, "total": len(cities_data)})


@api_view(["GET"])
def get_categories(request):
    return Response({
        "categories": [
            {"key": key, "label": val["label"]}
            for key, val in SERVICE_CATEGORY_MAP.items()
        ]
    })